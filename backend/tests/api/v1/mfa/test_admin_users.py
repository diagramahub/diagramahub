"""
Tests for the MFA admin user listing endpoint.
"""

from io import BytesIO

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from app.api.v1.ai_providers.schemas import (
    AIProviderConfig,
    AIProviderType,
    UserAISettingsInDB,
)
from app.api.v1.projects.schemas import ProjectInDB
from app.api.v1.users.schemas import UserInDB
from tests.utils import generate_test_password


@pytest.mark.integration
class TestAdminUsersListing:
    """Test suite for admin user listing details."""

    @pytest.mark.asyncio
    async def test_admin_user_list_includes_last_login_and_ai_models(
        self,
        authenticated_client: AsyncClient,
        client: AsyncClient,
    ):
        """Admin listing should expose login time and connected AI models."""
        email = "mfa-admin-list@example.com"
        password = generate_test_password()

        register_response = await client.post(
            "/api/v1/users/register",
            json={
                "email": email,
                "password": password,
                "full_name": "MFA Admin List",
            },
        )
        assert register_response.status_code == 201

        login_response = await client.post(
            "/api/v1/users/login",
            json={"email": email, "password": password},
        )
        assert login_response.status_code == 200

        user = await UserInDB.find_one(UserInDB.email == email)
        assert user is not None
        assert user.last_login_at is not None

        await ProjectInDB(
            name="Admin project",
            user_id=str(user.id),
        ).insert()

        await UserAISettingsInDB(
            user_id=str(user.id),
            providers=[
                AIProviderConfig(
                    provider=AIProviderType.GEMINI,
                    api_key="encrypted-key-placeholder",
                    model="gemini-2.5-flash",
                    is_active=True,
                    is_default=True,
                    parameters={},
                    display_name="Gemini",
                )
            ],
            auto_generate_on_save=False,
            default_provider=AIProviderType.GEMINI,
        ).insert()

        response = await authenticated_client.get(
            "/api/v1/mfa/admin/users",
            params={"search": email},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["items"]

        item = data["items"][0]
        assert item["email"] == email
        assert item["project_count"] == 1
        assert item["diagram_count"] == 0
        assert item["last_login_at"] is not None
        assert len(item["connected_ai_models"]) == 1
        assert item["connected_ai_models"][0]["provider"] == "gemini"
        assert item["connected_ai_models"][0]["model"] == "gemini-2.5-flash"

        export_response = await authenticated_client.get(
            "/api/v1/mfa/admin/users/export",
            params={"search": email},
            headers={"Accept-Language": "en"},
        )

        assert export_response.status_code == 200
        workbook = load_workbook(BytesIO(export_response.content))
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        assert "License Usage" in headers
        assert sheet.cell(row=2, column=headers.index("License Usage") + 1).value == "1 project / 0 diagrams"
