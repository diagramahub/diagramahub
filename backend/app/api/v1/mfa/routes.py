"""
MFA API routes for multi-factor authentication management.

Endpoints cover TOTP and email MFA setup, verification during login,
recovery codes, method switching, and status queries.
"""
import logging
from typing import Annotated

from fastapi import APIRouter, Depends, Header, HTTPException, status
from jose import JWTError

from app.api.v1.mfa.repository import MfaRepository
from app.api.v1.mfa.schemas import (
    MfaDisableRequest,
    MfaEnableTotpRequest,
    MfaResendRequest,
    MfaResendResponse,
    MfaSetDefaultMethodRequest,
    MfaSetupTotpResponse,
    MfaStatusResponse,
    MfaSwitchMethodRequest,
    MfaVerifyEmailActivationRequest,
    MfaVerifyRequest,
    RecoveryCodesResponse,
)
from app.api.v1.mfa.services import MfaService
from app.api.v1.users.routes import get_current_user_email
from app.api.v1.users.schemas import UserInDB
from app.core.security import create_access_token, decode_mfa_temp_token

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/mfa", tags=["MFA"])


def _get_mfa_service() -> MfaService:
    """Dependency injection for MFA service."""
    return MfaService(MfaRepository())


def _extract_lang(accept_language: str = Header(default="es", alias="Accept-Language")) -> str:
    """Extract the preferred language from the Accept-Language header."""
    if "en" in accept_language.lower():
        return "en"
    return "es"


# ---------------------------------------------------------------------------
# Helper: build MFA email HTML (same template used in users/routes.py login)
# ---------------------------------------------------------------------------

def _build_mfa_email_html(code: str, lang: str = "es") -> str:
    """Return an HTML email template for the MFA verification code."""
    if lang == "en":
        title = "Verification code"
        body = (
            "Use the following code to complete your sign in. "
            "This code expires in <strong>10 minutes</strong>."
        )
        footer_note = "If you did not try to sign in, you can ignore this email."
        copyright_text = "&copy; DiagramaHub. All rights reserved."
    else:
        title = "Código de verificación"
        body = (
            "Usa el siguiente código para completar tu inicio de sesión. "
            "Este código expira en <strong>10 minutos</strong>."
        )
        footer_note = "Si no intentaste iniciar sesión, puedes ignorar este correo."
        copyright_text = "&copy; DiagramaHub. Todos los derechos reservados."

    return f"""\
<!DOCTYPE html>
<html lang="{lang}">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>{title}</title>
</head>
<body style="margin:0;padding:0;background-color:#f4f4f7;font-family:Arial,Helvetica,sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background-color:#f4f4f7;padding:40px 0;">
    <tr>
      <td align="center">
        <table role="presentation" width="560" cellpadding="0" cellspacing="0" style="background-color:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08);">
          <tr>
            <td style="background:linear-gradient(135deg,#7c3aed 0%,#a855f7 50%,#9333ea 100%);padding:28px 40px;text-align:center;">
              <h1 style="margin:0;color:#ffffff;font-size:22px;font-weight:700;">DiagramaHub</h1>
            </td>
          </tr>
          <tr>
            <td style="padding:36px 40px 20px;">
              <h2 style="margin:0 0 16px;color:#1a1a2e;font-size:20px;font-weight:600;">{title}</h2>
              <p style="margin:0 0 24px;color:#51545e;font-size:15px;line-height:1.6;">
                {body}
              </p>
              <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
                <tr>
                  <td align="center" style="padding:8px 0 28px;">
                    <span style="display:inline-block;background-color:#faf5ff;color:#7c3aed;font-size:32px;font-weight:700;letter-spacing:8px;padding:16px 32px;border-radius:8px;border:1px solid #e9d5ff;">
                      {code}
                    </span>
                  </td>
                </tr>
              </table>
              <p style="margin:0;color:#9b9ba5;font-size:13px;line-height:1.5;">
                {footer_note}
              </p>
            </td>
          </tr>
          <tr>
            <td style="padding:20px 40px 28px;border-top:1px solid #eaeaec;text-align:center;">
              <p style="margin:0;color:#9b9ba5;font-size:12px;">{copyright_text}</p>
            </td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>"""


async def _send_mfa_email(email: str, code: str, lang: str = "es") -> None:
    """Send an MFA verification code via the configured email vendor.

    Failures are logged but do not raise — the caller decides how to handle.
    """
    subject = (
        "Your MFA verification code — DiagramaHub"
        if lang == "en"
        else "Tu código de verificación MFA — DiagramaHub"
    )
    try:
        from app.api.v1.integrations.email_service import EmailService
        from app.api.v1.integrations.repository import IntegrationsRepository

        email_service = EmailService(IntegrationsRepository())
        vendor = await email_service.get_default_email_vendor()
        html_content = _build_mfa_email_html(code, lang)
        await vendor.send_email(to=email, subject=subject, html_content=html_content)
    except Exception:
        logger.warning("Failed to send MFA email code to %s", email)
        raise


async def _get_user_by_email(email: str) -> UserInDB:
    """Fetch a user by email or raise 404."""
    user = await UserInDB.find_one(UserInDB.email == email)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuario no encontrado",
        )
    return user


# ---------------------------------------------------------------------------
# Authenticated endpoints (Bearer JWT)
# ---------------------------------------------------------------------------


@router.post("/setup-totp", response_model=MfaSetupTotpResponse)
async def setup_totp(
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
) -> MfaSetupTotpResponse:
    """Generate a TOTP secret and QR code for the authenticated user.

    The user must verify a code via ``/mfa/enable-totp`` to complete activation.
    """
    user = await _get_user_by_email(current_user_email)
    result = await mfa_service.setup_totp(str(user.id), user.email)
    return MfaSetupTotpResponse(**result)


@router.post("/enable-totp", response_model=None)
async def enable_totp(
    request: MfaEnableTotpRequest,
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
) -> dict:
    """Activate TOTP MFA by verifying a code from the authenticator app.

    Returns recovery codes on success, or a success message if codes already exist.
    """
    user = await _get_user_by_email(current_user_email)

    # The secret must be passed from the setup step. Since the setup step
    # returns the secret to the client and it is not yet persisted, the
    # client must send it back.  However, the current schema only carries
    # the verification code.  To keep the flow working without schema
    # changes, we re-generate the secret from the stored encrypted value
    # if TOTP was already set up, or we require the client to call
    # setup-totp first and use the secret from that response.
    #
    # For the enable flow, the service's enable_totp expects the plain
    # secret.  Since setup_totp doesn't persist the secret yet, the
    # frontend must hold it in memory and we need it here.  The design
    # has the frontend call setup-totp, get the secret, then call
    # enable-totp with the code.  We need the secret from the setup step.
    #
    # The pragmatic approach: re-generate a fresh setup and verify in one
    # shot is not ideal.  Instead, we store the secret temporarily in the
    # user's totp_secret_encrypted field during setup (but don't mark TOTP
    # as enabled).  Let's check if there's already an encrypted secret.
    from app.core.security import decrypt_totp_secret

    if not user.totp_secret_encrypted:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Primero debe ejecutar /mfa/setup-totp para generar el secreto TOTP",
        )

    try:
        secret = decrypt_totp_secret(user.totp_secret_encrypted)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Secreto TOTP inválido. Ejecute /mfa/setup-totp nuevamente",
        )

    result = await mfa_service.enable_totp(
        str(user.id), request.code, secret, set_as_default=request.set_as_default
    )
    if result["recovery_codes"] is None:
        return {"message": "TOTP MFA activado exitosamente", "codes": None}
    return {"codes": result["recovery_codes"]}


@router.post("/enable-email")
async def enable_email_mfa(
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
    lang: Annotated[str, Depends(_extract_lang)],
) -> dict:
    """Initiate email MFA activation by sending a verification code."""
    user = await _get_user_by_email(current_user_email)
    result = await mfa_service.enable_email_mfa(str(user.id), user.email)

    # Send the code via email
    plain_code = result["code"]
    await _send_mfa_email(user.email, plain_code, lang)

    return {"message": "Código de verificación enviado al correo electrónico"}


@router.post("/verify-email-activation", response_model=None)
async def verify_email_activation(
    request: MfaVerifyEmailActivationRequest,
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
) -> dict:
    """Confirm email MFA activation with the verification code.

    Returns recovery codes on success, or a success message if codes already exist.
    """
    user = await _get_user_by_email(current_user_email)
    result = await mfa_service.verify_email_activation(str(user.id), request.code)
    if result["recovery_codes"] is None:
        return {"message": "Email MFA activado exitosamente", "codes": None}
    return {"codes": result["recovery_codes"]}


@router.post("/disable")
async def disable_mfa(
    request: MfaDisableRequest,
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
) -> dict:
    """Disable a specific MFA method. Requires password confirmation."""
    user = await _get_user_by_email(current_user_email)
    await mfa_service.disable_mfa(str(user.id), request.password, request.method)
    return {"message": f"Método MFA '{request.method}' desactivado exitosamente"}


@router.get("/status", response_model=MfaStatusResponse)
async def get_mfa_status(
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
) -> MfaStatusResponse:
    """Return the current MFA status for the authenticated user."""
    user = await _get_user_by_email(current_user_email)
    result = await mfa_service.get_mfa_status(str(user.id))
    return MfaStatusResponse(**result)


@router.post("/regenerate-recovery-codes", response_model=RecoveryCodesResponse)
async def regenerate_recovery_codes(
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
) -> RecoveryCodesResponse:
    """Regenerate recovery codes, invalidating all previous ones."""
    user = await _get_user_by_email(current_user_email)

    if not user.mfa_enabled:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="MFA no está habilitado",
        )

    codes = await mfa_service.regenerate_recovery_codes(str(user.id))
    return RecoveryCodesResponse(codes=codes)


@router.put("/default-method")
async def set_default_method(
    request: MfaSetDefaultMethodRequest,
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
) -> dict:
    """Change the default MFA method for the authenticated user."""
    user = await _get_user_by_email(current_user_email)

    if not user.mfa_enabled:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="MFA no está habilitado",
        )

    await mfa_service.set_default_method(str(user.id), request.method)
    return {"message": f"Método MFA predeterminado cambiado a '{request.method}'"}


# ---------------------------------------------------------------------------
# Unauthenticated endpoints (use mfa_token)
# ---------------------------------------------------------------------------


@router.post("/verify")
async def verify_mfa(
    request: MfaVerifyRequest,
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
) -> dict:
    """Verify an MFA code during the login flow.

    Accepts TOTP codes, email codes, or recovery codes.
    Returns a full access token on success.
    """
    # Decode the temporary MFA token
    try:
        payload = decode_mfa_temp_token(request.mfa_token)
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token de verificación MFA inválido o expirado. Inicie sesión nuevamente",
        )

    email: str = payload.get("sub", "")
    attempt_count: int = payload.get("attempt_count", 0)
    available_methods: list[str] = payload.get("available_methods", [])

    # Check attempt limit (max 5)
    if attempt_count >= 5:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Máximo de intentos alcanzado. Inicie sesión nuevamente",
        )

    # Get user from DB
    user = await _get_user_by_email(email)
    user_id = str(user.id)

    # Determine the method to use
    method = request.method or payload.get("mfa_default_method")

    is_valid = False
    recovery_warning = None

    if request.is_recovery_code:
        is_valid = await mfa_service.verify_recovery_code(user_id, request.code)
        if is_valid:
            # Check remaining recovery codes
            mfa_status = await mfa_service.get_mfa_status(user_id)
            remaining = mfa_status.get("recovery_codes_remaining", 0)
            if remaining == 0:
                recovery_warning = (
                    "Has utilizado tu último código de recuperación. "
                    "Te recomendamos generar nuevos códigos desde la configuración de seguridad."
                )
    else:
        if method not in available_methods:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El método MFA solicitado no está activo",
            )
        is_valid = await mfa_service.verify_mfa_code(user_id, request.code, method)

    if not is_valid:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Código MFA inválido",
        )

    # Issue full access token with 5-day expiry (MFA enabled)
    access_token = create_access_token(email, mfa_enabled=True)

    response: dict = {
        "access_token": access_token,
        "token_type": "bearer",
    }
    if recovery_warning:
        response["recovery_warning"] = recovery_warning

    return response


@router.post("/switch-method")
async def switch_method(
    request: MfaSwitchMethodRequest,
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
    lang: Annotated[str, Depends(_extract_lang)],
) -> dict:
    """Switch to an alternative MFA method during login verification.

    If switching to email, sends a new verification code.
    """
    # Decode the temporary MFA token
    try:
        payload = decode_mfa_temp_token(request.mfa_token)
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token de verificación MFA inválido o expirado. Inicie sesión nuevamente",
        )

    email: str = payload.get("sub", "")
    available_methods: list[str] = payload.get("available_methods", [])

    # Verify the requested method is available
    if request.method not in available_methods:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El método MFA solicitado no está disponible",
        )

    # Get user from DB
    user = await _get_user_by_email(email)
    user_id = str(user.id)

    result = await mfa_service.switch_method(user_id, request.method, user.email)

    # If switching to email, send the code
    if request.method == "email" and "code" in result:
        await _send_mfa_email(user.email, result["code"], lang)

    return {"message": f"Método MFA cambiado a '{request.method}'"}


@router.post("/resend-email-code", response_model=MfaResendResponse)
async def resend_email_code(
    request: MfaResendRequest,
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
    lang: Annotated[str, Depends(_extract_lang)],
) -> MfaResendResponse:
    """Resend the email MFA verification code during login.

    Enforces resend limits (max 3) and cooldown (60s).
    """
    # Decode the temporary MFA token
    try:
        payload = decode_mfa_temp_token(request.mfa_token)
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token de verificación MFA inválido o expirado. Inicie sesión nuevamente",
        )

    email: str = payload.get("sub", "")

    # Get user from DB
    user = await _get_user_by_email(email)
    user_id = str(user.id)

    result = await mfa_service.resend_email_code(user_id, user.email)

    # Send the new code via email
    plain_code = result["code"]
    await _send_mfa_email(user.email, plain_code, lang)

    return MfaResendResponse(
        message="Código de verificación reenviado",
        resends_remaining=result["resends_remaining"],
    )


# ---------------------------------------------------------------------------
# Admin endpoints (requires admin role)
# ---------------------------------------------------------------------------


async def _require_admin(email: str) -> UserInDB:
    """Verify the current user is an admin. Raises 403 if not."""
    user = await UserInDB.find_one(UserInDB.email == email)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")
    if user.role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Acceso restringido a administradores")
    return user


@router.get("/admin/users")
async def admin_list_users(
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    page: int = 1,
    page_size: int = 20,
    search: str = "",
) -> dict:
    """List users with MFA status. Admin only. Supports pagination and search."""
    await _require_admin(current_user_email)

    # Build query
    query = UserInDB.find()
    if search:
        import re
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query = UserInDB.find({"$or": [{"email": pattern}, {"full_name": pattern}]})

    total = await query.count()

    skip = (page - 1) * page_size
    users = await query.skip(skip).limit(page_size).sort("-created_at").to_list()

    items = []
    for u in users:
        unused_codes = sum(
            1 for c in u.recovery_codes
            if not (c.used if hasattr(c, "used") else c.get("used", False))
        )

        # Fetch subscription & plan name
        plan_name = None
        try:
            from app.api.v1.subscriptions.schemas import SubscriptionInDB, PlanInDB
            sub = await SubscriptionInDB.find_one(
                SubscriptionInDB.user_id == str(u.id),
                SubscriptionInDB.status == "active",
            )
            if sub:
                plan = await PlanInDB.get(sub.plan_id)
                if plan:
                    plan_name = plan.name
        except Exception:
            pass

        items.append({
            "id": str(u.id),
            "email": u.email,
            "full_name": u.full_name,
            "role": u.role,
            "is_active": u.is_active,
            "mfa_enabled": u.mfa_enabled,
            "mfa_methods": u.mfa_methods,
            "mfa_default_method": u.mfa_default_method,
            "recovery_codes_remaining": unused_codes,
            "created_at": u.created_at.isoformat() if u.created_at else None,
            "plan_name": plan_name,
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
    }


@router.post("/admin/users/{user_id}/reset-mfa")
async def admin_reset_user_mfa(
    user_id: str,
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    mfa_service: Annotated[MfaService, Depends(_get_mfa_service)],
) -> dict:
    """Reset (disable) all MFA methods for a user. Admin only."""
    admin = await _require_admin(current_user_email)

    from bson import ObjectId

    target_user = await UserInDB.get(ObjectId(user_id))
    if not target_user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")

    if not target_user.mfa_enabled:
        return {"message": "El usuario no tiene MFA habilitado"}

    # Disable all methods
    for method in list(target_user.mfa_methods):
        await mfa_service.repository.disable_mfa(user_id, method)

    logger.info(
        "Admin %s reset MFA for user %s (%s)",
        admin.email,
        target_user.email,
        user_id,
    )

    return {"message": f"MFA desactivado para {target_user.email}"}


@router.get("/admin/users/export")
async def admin_export_users_excel(
    current_user_email: Annotated[str, Depends(get_current_user_email)],
    lang: Annotated[str, Depends(_extract_lang)],
):
    """Export all users to an Excel file. Admin only."""
    await _require_admin(current_user_email)

    import io
    from datetime import datetime

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    users = await UserInDB.find_all().sort("-created_at").to_list()

    wb = Workbook()
    ws = wb.active
    ws.title = "Users" if lang == "en" else "Usuarios"

    # Headers
    if lang == "en":
        headers = [
            "Email", "Full Name", "Role", "Plan", "Active", "MFA Enabled",
            "MFA Methods", "Default Method", "Recovery Codes Remaining",
            "Created At",
        ]
    else:
        headers = [
            "Correo", "Nombre completo", "Rol", "Plan", "Activo", "MFA Habilitado",
            "Métodos MFA", "Método predeterminado", "Códigos de recuperación",
            "Fecha de registro",
        ]

    # Style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for row_idx, u in enumerate(users, 2):
        unused_codes = sum(
            1 for c in u.recovery_codes
            if not (c.used if hasattr(c, "used") else c.get("used", False))
        )
        methods_str = ", ".join(
            ("Email" if m == "email" else "TOTP") for m in u.mfa_methods
        )
        default_str = ""
        if u.mfa_default_method:
            default_str = "Email" if u.mfa_default_method == "email" else "TOTP"
        created_str = u.created_at.strftime("%Y-%m-%d %H:%M") if u.created_at else ""

        # Fetch plan name
        plan_name = ""
        try:
            from app.api.v1.subscriptions.schemas import SubscriptionInDB, PlanInDB
            sub = await SubscriptionInDB.find_one(
                SubscriptionInDB.user_id == str(u.id),
                SubscriptionInDB.status == "active",
            )
            if sub:
                plan = await PlanInDB.get(sub.plan_id)
                if plan:
                    plan_name = plan.name
        except Exception:
            pass

        yes = "Yes" if lang == "en" else "Sí"
        no = "No"

        row_data = [
            u.email,
            u.full_name or "",
            u.role,
            plan_name,
            yes if u.is_active else no,
            yes if u.mfa_enabled else no,
            methods_str,
            default_str,
            unused_codes,
            created_str,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M")
    filename = f"diagramahub_users_{timestamp}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
